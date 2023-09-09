<<<<<<< HEAD
import openpyxl

# La función leer_Patron se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el patrón y de salida se da una matriz.
def leerPatron(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizPatron = [[None for _ in range(5)] for _ in range(5)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizPatron, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizPatron, Error)

    # Se estudia el estado del patron dado por el usuario
    for fila in range(1, 6):
        for columna in range(1, 6):
            celda = sheet.cell(row=fila, column=columna).value
            if celda == 100:
                matrizPatron[fila-1][columna-1] = 100
            elif celda == 200:
                matrizPatron[fila-1][columna-1] = 200
            elif celda == 300:
                matrizPatron[fila-1][columna-1] = 300
            else:
                matrizPatron[fila-1][columna-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    imprimir_matriz(matrizPatron)
    return (matrizPatron, Error)

# La función leerSuministro se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el suministro y de salida se da una matriz.
def leerSuministro(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizSuministro = [[None for _ in range(5)] for _ in range(5)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizSuministro, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizSuministro, Error)

    # Se estudia el estado del patron dado por el usuario
    for fila in range(1, 6):
        for columna in range(1, 6):
            celda = sheet.cell(row=fila, column=columna).value
            if celda == 100:
                matrizSuministro[fila-1][columna-1] = 100
            elif celda == 200:
                matrizSuministro[fila-1][columna-1] = 200
            elif celda == 300:
                matrizSuministro[fila-1][columna-1] = 300
            else:
                matrizSuministro[fila-1][columna-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    return (matrizSuministro)

def imprimir_matriz(matriz):
    print ("\nIMPRIMIENNDO MATRIZ:\n")
    for fila in matriz:
        print(fila)

archivoPatron = "archivoPatron.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
#archivoSuministro = "archivoSuministro.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
matrizPatron, Error = leerPatron(archivoPatron)
#matrizSuministro = leerSuministro(archivoSuministro)
#imprimir_matriz(matrizPatron)
#imprimir_matriz(matrizSuministro)
=======
import openpyxl

# La función leer_Patron se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el patrón y de salida se da una matriz.
def leerPatron(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizPatron = [[None for _ in range(5)] for _ in range(5)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizPatron, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizPatron, Error)

    # Se estudia el estado del patron dado por el usuario
    for fila in range(1, 6):
        for columna in range(1, 6):
            celda = sheet.cell(row=fila, column=columna).value
            if celda == "A":
                matrizPatron[fila-1][columna-1] = "A"
            elif celda == "B":
                matrizPatron[fila-1][columna-1] = "B"
            elif celda == "C":
                matrizPatron[fila-1][columna-1] = "C"
            else:
                matrizPatron[fila-1][columna-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    imprimir_matriz(matrizPatron)
    return (matrizPatron, Error)

# La función leerSuministro se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el suministro y de salida se da una matriz.
def leerSuministro(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizSuministro = [[None for _ in range(5)] for _ in range(5)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizSuministro, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizSuministro, Error)

    # Se estudia el estado del patron dado por el usuario
    for fila in range(1, 6):
        for columna in range(1, 6):
            celda = sheet.cell(row=fila, column=columna).value
            if celda == "A":
                matrizSuministro[fila-1][columna-1] = "A"
            elif celda == "B":
                matrizSuministro[fila-1][columna-1] = "B"
            elif celda == "C":
                matrizSuministro[fila-1][columna-1] = "C"
            else:
                matrizSuministro[fila-1][columna-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    return (matrizSuministro)

def imprimir_matriz(matriz):
    print ("\nIMPRIMIENNDO MATRIZ:\n")
    for fila in matriz:
        print(fila)

#archivoPatron = "archivoPatron.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
#archivoSuministro = "archivoSuministro.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
#matrizPatron, Error = leerPatron(archivoPatron)
#matrizSuministro = leerSuministro(archivoSuministro)
#imprimir_matriz(matrizPatron)
#imprimir_matriz(matrizSuministro)

>>>>>>> 1dd20a62434644c76f75215e51fbfe1d238391f4
