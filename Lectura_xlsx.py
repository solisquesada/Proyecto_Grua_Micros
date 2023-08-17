import openpyxl

# La función leer_Patron se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el patrón y de salida se da una matriz.
def leerPatron(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizPatron = [[None for _ in range(5)] for _ in range(5)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizPatron, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizPatron, Error)

    # Se estudia el estado del patron dado por el usuario
    for fila in range(1, 6):
        for columna in range(1, 6):
            celda = sheet.cell(row=fila, column=columna).value
            if celda == "A":
                matrizPatron[fila-1][columna-1] = "A"
            elif celda == "B":
                matrizPatron[fila-1][columna-1] = "B"
            elif celda == "C":
                matrizPatron[fila-1][columna-1] = "C"
            else:
                matrizPatron[fila-1][columna-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    return (matrizPatron, Error)

# La función leerSuministro se encarga de obtener información
# del excel. Como entrada se da el archivo donde el usuario
# ingresa el suministro y de salida se da una matriz.
def leerSuministro(archivo):
    Error = 0 # Se empieza el error en 0
    # Se definen las matrizes a llenar
    matrizPatron = [None for _ in range(25)]

    try:
        # Se abre el archivo .xlsx
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
    except FileNotFoundError:
        # En caso de no encontrarse el archivo se avisa al usuario
        Error = 101 # Se define el error 101 como error por no archivo
        print(f"El archivo '{archivo}' no fue encontrado.")
        return (matrizPatron, Error)
    except PermissionError:
        # En caso de no tener permisos al archivo avisa al usuario
        Error = 102 # Se define el error 102 como error por archivo bloqueado
        print(f"El archivo '{archivo}' está bloqueado.")
        return (matrizPatron, Error)

    # Se escribe la información en una lista (se llama fila porque es la fila
    # de la matriz, no del archivo)
    for fila in range(1, 25):
        celda = sheet.cell(row=2, column=fila).value
        if celda == "A":
            matrizPatron[fila-1] = "A"
        elif celda == "B":
            matrizPatron[fila-1] = "B"
        elif celda == "C":
            matrizPatron[fila-1] = "C"
        else:
            matrizPatron[fila-1] = None

    # Para que esta función funcione, el archivo tiene que tener el formato dado
    # por nosotros en la plantilla de excel.
    return (matrizPatron, Error)

def imprimir_matriz(matriz):
    print ("\nIMPRIMIENNDO MATRIZ:\n")
    for fila in matriz:
        print(fila)

archivoPatron = "archivoPatron.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
archivoSuministro = "archivoSuministro.xlsx"  # Reemplaza con la ruta a tu archivo XLSX o CSV
matrizPatron, Error = leerPatron(archivoPatron)
matrizSuministro, Error = leerSuministro(archivoSuministro)
imprimir_matriz(matrizPatron)
imprimir_matriz(matrizSuministro)

